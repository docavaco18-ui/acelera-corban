"""Router PowerHub — enriquecimento de telefone via CPF (/api/powerhub/*)."""
import io
import json
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ..auth_deps import require_user, AuthUser
from ..database import db as get_db
from ..db_scoped import scoped
from ..banks.powerhub.credentials_helper import get_powerhub_runtime_creds
from ..services import powerhub_upload_jobs, powerhub_bot_service

router = APIRouter(prefix="/api/powerhub", tags=["powerhub"])
_events: list[dict] = []


def _on_event(event: dict):
    _events.append(event)
    if len(_events) > 500:
        _events.pop(0)


# ─── Upload ──────────────────────────────────────────────────────────────────

@router.post("/leads/upload", status_code=202)
async def upload_csv(
    file: UploadFile = File(...),
    user: AuthUser = Depends(require_user),
):
    content = await file.read()
    if not content:
        raise HTTPException(400, "Arquivo vazio")
    return await powerhub_upload_jobs.start_upload(content, user.user_id, file_name=file.filename)


@router.get("/leads/upload/{job_id}")
async def upload_status(job_id: str, _: AuthUser = Depends(require_user)):
    state = await powerhub_upload_jobs.get_job(job_id)
    if state is None:
        raise HTTPException(404, "Job não encontrado")
    return state


# ─── Leads ───────────────────────────────────────────────────────────────────

@router.get("/leads/")
async def list_leads(
    status: str = None,
    batch_id: str | None = None,
    page: int = 1,
    limit: int = 50,
    user: AuthUser = Depends(require_user),
):
    db = get_db()
    q = scoped(db, "powerhub_leads", user.user_id).select("*").order("created_at", desc=True)
    if status:
        q = q.eq("status", status)
    if batch_id:
        q = q.eq("batch_id", batch_id)
    result = q.range((page - 1) * limit, page * limit - 1).execute()
    return {"data": result.data, "page": page}


def _fmt_phone(raw: str) -> str:
    digits = "".join(c for c in str(raw) if c.isdigit())
    if not digits:
        return ""
    if digits.startswith("55") and len(digits) >= 12:
        return f"+{digits}"
    return f"+55{digits}"


@router.get("/leads/export")
async def export_xlsx(
    batch_id: str | None = None,
    status: str = "found",
    user: AuthUser = Depends(require_user),
):
    """Exporta Excel com 2 abas: Enriquecido (PowerHub) + Original (telefone do CSV)."""
    db = get_db()
    q = scoped(db, "powerhub_leads", user.user_id).select("*")
    if status:
        q = q.eq("status", status)
    if batch_id:
        q = q.eq("batch_id", batch_id)
    # Cap export em 50k linhas pra evitar OOM
    leads = q.limit(50_000).execute().data or []

    wb = openpyxl.Workbook()

    # ── Aba 1: Enriquecido (PowerHub) ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Enriquecido"
    hdr_fill = PatternFill("solid", fgColor="7C3AED")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(["CPF", "Nome", "Telefone", "Fonte"], 1):
        cell = ws1.cell(row=1, column=col, value=title)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for lead in leads:
        cpf = lead.get("cpf", "")
        nome = lead.get("nome", "") or ""
        phones_raw = lead.get("phones") or "[]"
        try:
            phones = json.loads(phones_raw) if isinstance(phones_raw, str) else (phones_raw or [])
        except Exception:
            phones = []

        if phones:
            for phone in phones:
                ws1.append([cpf, nome, _fmt_phone(phone), "PowerHub"])
                row += 1
        else:
            ws1.append([cpf, nome, "", "PowerHub"])
            row += 1

    for col in [1, 2, 3, 4]:
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = [16, 35, 18, 12][col - 1]

    # ── Aba 2: Original (telefone que veio no CSV) ─────────────────────────────
    has_original = any(lead.get("telefone_original") for lead in leads)
    if has_original:
        ws2 = wb.create_sheet("Original")
        hdr_fill2 = PatternFill("solid", fgColor="1D4ED8")
        for col, title in enumerate(["CPF", "Nome", "Telefone Original"], 1):
            cell = ws2.cell(row=1, column=col, value=title)
            cell.fill = hdr_fill2
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for lead in leads:
            fone = lead.get("telefone_original")
            ws2.append([lead.get("cpf", ""), lead.get("nome", "") or "", _fmt_phone(fone) if fone else ""])
        for col in [1, 2, 3]:
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = [16, 35, 18][col - 1]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=powerhub_telefones.xlsx"},
    )


# ─── Batches ─────────────────────────────────────────────────────────────────

@router.get("/batches/")
async def list_batches(user: AuthUser = Depends(require_user)):
    db = get_db()
    rows = scoped(db, "powerhub_batches", user.user_id).select("*").order("created_at", desc=True).execute().data or []
    return {"data": rows}


@router.get("/batches/current")
async def current_batch(user: AuthUser = Depends(require_user)):
    db = get_db()
    rows = (
        scoped(db, "powerhub_batches", user.user_id)
        .select("*").eq("status", "active").order("created_at", desc=True).limit(1).execute().data or []
    )
    return rows[0] if rows else None


# ─── Bot ─────────────────────────────────────────────────────────────────────

@router.post("/bot/start")
async def bot_start(
    request: Request,
    num_workers: int = 3,
    batch_id: str | None = None,
    user: AuthUser = Depends(require_user),
):
    db = get_db()
    creds = get_powerhub_runtime_creds(user.user_id, db)
    pool = request.app.state.powerhub_pool
    result = await powerhub_bot_service.start_bot(
        pool=pool,
        user_id=user.user_id,
        creds=creds,
        db=db,
        num_workers=num_workers,
        batch_id=batch_id,
        on_event=_on_event,
    )
    return result


@router.post("/bot/stop")
async def bot_stop(request: Request, user: AuthUser = Depends(require_user)):
    pool = request.app.state.powerhub_pool
    await pool.stop(user.user_id)
    return {"ok": True}


@router.get("/bot/status")
async def bot_status(request: Request, user: AuthUser = Depends(require_user)):
    pool = request.app.state.powerhub_pool
    handle = pool.status(user.user_id)
    if not handle:
        return {"status": "idle"}
    return {
        "status": "running",
        "run_id": handle.run_id,
        "num_workers": handle.num_workers,
        "started_at": handle.started_at.isoformat(),
    }


@router.get("/bot/events")
async def bot_events(_: AuthUser = Depends(require_user)):
    return {"data": _events[-100:]}


# ─── Stats ───────────────────────────────────────────────────────────────────

@router.get("/stats")
async def stats(batch_id: str | None = None, user: AuthUser = Depends(require_user)):
    db = get_db()
    q = scoped(db, "powerhub_leads", user.user_id).select("status,phones")
    if batch_id:
        q = q.eq("batch_id", batch_id)
    leads = q.execute().data or []

    counts = {"pending": 0, "processing": 0, "found": 0, "not_found": 0, "error": 0}
    total_phones = 0

    for lead in leads:
        s = lead.get("status", "pending")
        counts[s] = counts.get(s, 0) + 1
        try:
            phones = json.loads(lead.get("phones") or "[]")
            total_phones += len(phones)
        except Exception:
            pass

    return {
        **counts,
        "total": len(leads),
        "total_phones": total_phones,
    }
